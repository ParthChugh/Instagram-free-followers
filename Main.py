import pyautogui, sys, os
import webbrowser
import time
import names,string,random
from openpyxl import load_workbook

    
def random_char(y):
    return ''.join(random.choice(string.ascii_letters) for x in range(y))

chrome_path = 'C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s --incognito'
for i in range(1,2):
    from openpyxl import load_workbook
    wb = load_workbook('Names.xlsx')
    ws = wb.active

    os.system("powershell -C Start-Process chrome.exe -ArgumentList @( '-incognito', 'https://www.instagram.com/?hl=en' )")
    time.sleep(2)
    pyautogui.click(x=372, y=241)
    pyautogui.hotkey('win','up')
    time.sleep(14)
    pyautogui.moveRel(0, 100)
    pyautogui.click(x=838, y=401)

    name = names.get_first_name()
    last = names.get_last_name()

    pyautogui.typewrite(random_char(5)+'@'+str(random_char(5))+ '.com')
    pyautogui.click(x=838, y=457)
    pyautogui.typewrite(name)
    
    pyautogui.moveRel(0, 50)
    pyautogui.click(x=838, y=491)

    username = name+'_'+last+'_'+str(random.randint(400,1956))
    pyautogui.typewrite(username)
    pyautogui.click(x=838, y=537)
    
    password = 'OsheenBot'
    pyautogui.typewrite(password)

    ws.append([username,password])
    wb.save("Names.xlsx")

    pyautogui.click(x=838, y=587)
    time.sleep(4)
    os.system("powershell -C Start-Process chrome.exe -ArgumentList @( '-incognito', 'https://www.instagram.com/travelandfood0055/?hl=en' )")
    time.sleep(4)
    pyautogui.click(x=840, y=215)
    time.sleep(4)
    os.system("taskkill /im chrome.exe /f")
    
