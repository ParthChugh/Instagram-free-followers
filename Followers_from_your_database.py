from openpyxl import load_workbook
import os
import pyautogui, sys
import time

wb = load_workbook('Names.xlsx')
ws = wb.get_sheet_by_name('Sheet')
#ws['A150'].value

#ws = wb['Sheet']

a = ws.max_row

for i in range(2,a):
    username = ws['A'+str(i)].value 
    password = ws['B'+str(i)].value
    
    os.system("powershell -C Start-Process chrome.exe -ArgumentList @( '-incognito', 'https://www.instagram.com/accounts/login/' )")
    time.sleep(5)
    pyautogui.click(x=372, y=241)
    pyautogui.hotkey('win','up')
    time.sleep(1)
    pyautogui.click(x=595, y=231)
    pyautogui.typewrite(username)
    pyautogui.click(x=595, y=272)
    pyautogui.typewrite(password)
    pyautogui.click(x=595, y=313)
    time.sleep(2)
    os.system("powershell -C Start-Process chrome.exe -ArgumentList @( '-incognito', 'https://www.instagram.com/jerrod_miller_749/?hl=en' )")
    time.sleep(2)
    pyautogui.click(x=775, y=220)
    time.sleep(4)
    os.system("taskkill /im chrome.exe /f")
    
    
        
    
    
